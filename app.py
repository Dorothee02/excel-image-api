import os
import tempfile
import zipfile
from flask import Flask, request, jsonify, send_file
import shutil
import pandas as pd
import openpyxl
import xml.etree.ElementTree as ET
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

app = Flask(__name__)

# 对应的 XML 命名空间
NS = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a':   'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r':   'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
}

REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
ns = { 'rel': REL_NS }

def obtainJanValueColumnIdx(wb):
    jan_column = None
    jan_header_row = None

    for sheet in wb.worksheets:
        for row_idx in range(1, min(20, sheet.max_row + 1)):  # 搜尋前20行
            for col_idx in range(1, min(20, sheet.max_column + 1)):  # 搜尋前20列
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value and isinstance(cell_value, str) and "JAN" in cell_value:
                    jan_column = col_idx
                    jan_header_row = row_idx
                    break
            if jan_column:
                break
        if jan_column:
            break
    return jan_column, jan_header_row

def map_images_to_cells(xlsx_path, jan_column):
    wb = load_workbook(xlsx_path, data_only=True)

    with zipfile.ZipFile(xlsx_path, 'r') as zf:
        all_files = zf.namelist()
        results = []

        for idx, sheet in enumerate(wb.worksheets, start=1):
            rels_path = f'xl/worksheets/_rels/sheet{idx}.xml.rels'
            if rels_path not in all_files:
                continue
            rels_xml = ET.fromstring(zf.read(rels_path))

            alist = rels_xml.findall('rel:Relationship', ns)
            for rel in alist:
                if rel.attrib.get('Type','').endswith('/drawing'):
                    drawing_target = rel.attrib['Target']
                    drawing_path = 'xl/' + drawing_target.replace('../','')
                    break
            else:
                continue

            drawing_xml = ET.fromstring(zf.read(drawing_path))
            drawing_rels = f'xl/drawings/_rels/{drawing_path.split("/")[-1]}.rels'
            rels2_xml = ET.fromstring(zf.read(drawing_rels))

            blist = rels2_xml.findall('rel:Relationship', ns)
            rid2media = {
                r.attrib['Id']: 'xl/' + r.attrib['Target'].replace('../','')
                for r in blist
            }

            for anchor in drawing_xml.findall('.//xdr:twoCellAnchor', NS) + drawing_xml.findall('.//xdr:oneCellAnchor', NS):
                frm = anchor.find('xdr:from', NS)
                row = int(frm.find('xdr:row', NS).text) + 1
                col = int(frm.find('xdr:col', NS).text) + 1
                cell = get_column_letter(col) + str(row)

                blip = anchor.find('.//a:blip', NS)
                rId  = blip.attrib[f'{{{NS["r"]}}}embed']
                media_file = rid2media.get(rId)

                jan = sheet.cell(row=row, column=jan_column).value

                results.append({
                    'sheet':     sheet.title,
                    'cell':      cell,
                    'row':       row,
                    'col':       col,
                    'jan_value': jan or f'unknown_row{row}',
                    'media':     media_file,
                    'media_name': os.path.basename(media_file)
                })

    return results

def clear_folder(folder_path):
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)
    os.makedirs(folder_path, exist_ok=True)

def proceed(filepath, output_dir, tmpdir):
    wb = openpyxl.load_workbook(filepath)
    jan_column, jan_header_row = obtainJanValueColumnIdx(wb)
    
    if not jan_column:
        return None, None, []

    with zipfile.ZipFile(filepath, 'r') as zip_ref:
        zip_ref.extractall(tmpdir)

    source_img_folder = os.path.join(tmpdir, "xl/media")
    if not os.path.exists(source_img_folder):
        return None, None, []

    extracted_images = []
    mapping = map_images_to_cells(filepath, jan_column=jan_column)
    
    for info in mapping:
        jan_value = info['jan_value']
        media_name = info['media_name']
        position = info['cell']
        row = info['row']
        
        new_filename = f"{jan_value}_{media_name}"
        output_path = os.path.join(output_dir, new_filename)
        source_img_path = os.path.join(source_img_folder, media_name)
        
        if os.path.exists(source_img_path):
            shutil.copy2(source_img_path, output_path)
            extracted_images.append({
                "original_name": media_name,
                "position": position,
                "row": row,
                "jan_value": jan_value,
                "saved_as": new_filename
            })

    # 去除重複的檔案
    janId = []
    filter_extracted_images = []
    for img in extracted_images:
        if img['jan_value'] not in janId:
            janId.append(img['jan_value'])
            filter_extracted_images.append(img)

    # 建立ZIP檔案
    zip_output = os.path.join(tmpdir, "extracted_images.zip")
    with zipfile.ZipFile(zip_output, 'w') as zipf:
        for img_info in filter_extracted_images:
            saved_file = img_info["saved_as"]
            img_path = os.path.join(output_dir, saved_file)
            if os.path.exists(img_path):
                zipf.write(img_path, arcname=saved_file)

    return jan_column, jan_header_row, filter_extracted_images, zip_output

@app.route('/upload', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files['file']
    
    # 使用臨時目錄
    with tempfile.TemporaryDirectory() as temp_dir:
        # 建立臨時目錄和輸出目錄
        tmpdir = os.path.join(temp_dir, 'temp')
        output_dir = os.path.join(temp_dir, 'output')
        os.makedirs(tmpdir, exist_ok=True)
        os.makedirs(output_dir, exist_ok=True)
        
        # 保存上傳的文件
        filepath = os.path.join(tmpdir, file.filename)
        file.save(filepath)
        
        # 處理文件
        jan_column, jan_header_row, extracted_images, zip_output = proceed(filepath, output_dir, tmpdir)
        
        if not jan_column:
            return jsonify({"error": "Could not find JAN column in the Excel file"}), 400
        
        # 回傳 ZIP 檔案（如果需要）
        if os.path.exists(zip_output):
            return send_file(
                zip_output,
                mimetype='application/zip',
                as_attachment=True,
                download_name='extracted_images.zip'
            )
        else:
            # 如果只需要回傳 JSON 資訊
            return jsonify({
                "status": "success",
                "message": "Images extracted and renamed with JAN codes",
                "jan_column_found": f"Column {openpyxl.utils.get_column_letter(jan_column)} (Header at row {jan_header_row})",
                "extracted_images": extracted_images
            })

@app.route('/test', methods=['GET'])
def test():
    return jsonify({"status": "API is running"})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)
